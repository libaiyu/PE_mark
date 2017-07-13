
import openpyxl


def isleap(year):
    """Return True for leap years, False for non-leap years."""
    return year % 4 == 0 and (year % 100 != 0 or year % 400 == 0)


def adjust_date(month_v, date):
    '''  '''
    if month_v in Small_month:
        if date > 30:
            month_v += 1
            date -= 30
    elif month_v in Big_month:
        if date > 31:
            month_v += 1
            date -= 31
    else:
        if isleap(year):
            if date > 29:
                month_v += 1
                date -= 29            
        else:
            if date > 28:
                month_v += 1
                date -= 28
    return month_v, date


def adjust_mon(month_n):
    '''  '''
    global year
    if month_n > 12:
        year += 1
        month_n = 1
    return month_n


'''
def adjust_year(month_n):
    '''  '''
    if month_n > 12:
        year += 1
    return year
'''


def main():
    '''  '''
    global year, month, date_monday, date_sunday
    str_year = input('Please input the year:')
    year = int(str_year)
    date_monday = int(input('Please input the date of the first week:'))
    str_month = input('Please input the month of the first week:')
    month = int(str_month)
    wb = openpyxl.Workbook()
    sheet = wb.get_active_sheet()
    sheet.cell(row=1, column=1).value = '周次'   # Write week.
    sheet.cell(row=1, column=2).value = '日/月至日/月'     # Write week_date.
    sheet.cell(row=1, column=3).value = '年'     # Write week_date.
    for row in range(2, 45, 2):   # One semester has 22 weeks(at most).
        date_month = str(date_monday) + '/' + str(month)
        date_sunday = date_monday + 6
        month, date_sunday = adjust_date(month, date_sunday)
        month = adjust_mon(month)
        date_month += '-' + str(date_sunday) + '/' + str(month)   # Form the week_date.
        sheet.cell(row=row, column=1).value = row/2      # Write week.
        sheet.cell(row=row, column=2).value = date_month   # Write week_date.
        sheet.cell(row=row, column=3).value = year
        date_monday = date_sunday + 1
        month, date_monday = adjust_date(month, date_monday)
        month = adjust_mon(month)
    wb.save('calendar-'+str_year+'-'+str_month+'.xlsx')


if __name__ == '__main__':
    '''  '''
    print('---Begin--------')
    date_monday = 4
    month = 9
    Big_month = [10, 12, 1, 3, 5, 7, 8]
    Small_month = [9, 11, 4, 6]
    Feb = [2]
    main()
    print('---End--------')
