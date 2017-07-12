

''' make the score dict of the file(“ 高考体育成绩对照表”), from the file
(“高考体育成绩测量表”)to get the information, then look it up in the score dict
to get the score. Finally write the score to a new file.     2017-7-12 '''


import os
import re
import openpyxl
import logging
from get_item_row_column import get_item_row_column


groups_measure_score_man = {}
groups_measure_score_woman = {}
REG = re.compile(r'标准表')
GROUPS = ('甲', '乙', '丙', '丁')
ITEMS = ('100米', '跳远', '铅球', '８００米')


def get_measure_score(file, item_column, item_row):
    ''' get the measure value in the given file. '''
    measure_score_man = {}
    measure_score_woman = {}
    wb = openpyxl.load_workbook(file)
    sheet = wb.get_active_sheet()
    for row in range(item_row+2, sheet.max_row+1):
        man_measure = sheet.cell(row=row, column=item_column).value
        man_score = sheet.cell(row=row, column=item_column+1).value
        measure_score_man[man_measure] = man_score
        woman_measure = sheet.cell(row=row, column=item_column+2).value
        woman_score = sheet.cell(row=row, column=item_column+3).value
        measure_score_woman[woman_measure] = woman_score
    return measure_score_man, measure_score_woman


def get_items_position(file):
    ''' get the row and column of the ITEMS in the given file. '''
    global group_item_cols_rows
    items_cols_rows = {}
    for item in ITEMS:
        item_column, item_row = get_item_row_column(file, item)
        items_cols_rows[item] = (item_column, item_row)
    return items_cols_rows


def get_items_measure_score(file):
    '''   '''
    global groups_measure_score
    group = file[3]
    items_cols_rows = get_items_position(file)
    items_measure_score_man = {}
    items_measure_score_woman = {}
    for item, (item_column, item_row) in items_cols_rows.items():
        measure_score_man, measure_score_woman =\
                           get_measure_score(file, item_column, item_row)
        items_measure_score_man[item] = measure_score_man
        items_measure_score_woman[item] = measure_score_woman
    groups_measure_score_man[group] = items_measure_score_man
    groups_measure_score_woman[group] = items_measure_score_woman
    return items_measure_score_man, items_measure_score_woman


def get_measure_information(measure_file):
    '''先获取所需的信息：姓名，男女，项目，测量值，组别。'''
    wb = openpyxl.load_workbook(measure_file)
    sheet = wb.get_active_sheet()
    measure_information_mark = []
    col = 2
    for row in range(1, sheet.max_row+1):
        name = sheet.cell(row=row, column=col).value
        if name:
            group = sheet.cell(row=row, column=col-1).value
            gender = sheet.cell(row=row, column=col+1).value
            item_measure = []
            for num in range(len(ITEMS)):
                item = sheet.cell(row=2, column=col+2+num).value
                measure = sheet.cell(row=row, column=col+2+num).value
                item_measure.append([item, measure])
            measure_information_mark.append((group, name, gender, item_measure))
    for test in range(3):
        print(measure_information_mark[test])
    print()
    return measure_information_mark


def information_to_mark(measure_information):
    '''然后根据参数中提供的信息找到相应的成绩。'''
    for each_information in measure_information:
        group = each_information[0]
        gender = each_information[2]
        for num in range(len(ITEMS)):
            item = each_information[3][num][0]
            measure = each_information[3][num][1]
            if gender == '女':
                mark = groups_measure_score_woman[group][item][measure]
            elif gender == '男':
                mark = groups_measure_score_man[group][item][measure]
            each_information[3][num].append(mark)
    for test in range(3):
        print(measure_information[test])
    print()
    return measure_information


def write(measure_information_mark, new_file):
    ''' write in a new file. '''
    wb = openpyxl.Workbook()
    sheet = wb.get_active_sheet()
    row = 1
    for each_information in measure_information_mark:
        row += 1
        column = 2
        group, name, gender, item_measure = each_information
        sheet.cell(row=row, column=column).value = group
        sheet.cell(row=row, column=column+1).value = name
        sheet.cell(row=row, column=column+2).value = gender
        num = 0
        for item, measure, score in item_measure:
            sheet.cell(row=row, column=column+3+num).value = item
            sheet.cell(row=row, column=column+4+num).value = measure
            sheet.cell(row=row, column=column+5+num).value = score
            num += 3
    wb.save(new_file)


def test():
    '''  '''
    # get the file that include "标准表".
    filelist = []
    for file in os.listdir():
        if REG.search(file):
            filelist.append(file)
    for file in filelist:
        print(file)
        items_measure_score_man, items_measure_score_woman =\
                                 get_items_measure_score(file)
    # get measure information.
    measure_file = '高考体育成绩测量表.xlsx'
    measure_information = get_measure_information(measure_file)
    # get the score for the measure information.
    measure_information_mark = information_to_mark(measure_information)
    # write the information and score to a new file.
    new_file = '体育成绩表4.xlsx'
    write(measure_information_mark, new_file)


if __name__ == "__main__":
    '''  '''
    print('--------Start of program---------')
    test()
    print('-------End--------')


