
import openpyxl


def get_item_row_column(file, item_name):
    ''' Get the column number of the item. '''
    item_column = None    # item_name not found, item_column = None
    wb = openpyxl.load_workbook(file)
    sheet = wb.get_active_sheet()
    for row in range(1, sheet.max_row+1):
        for column in range(1, sheet.max_column+1):
            if sheet.cell(row=row, column=column).value == item_name:
                item_row = row
                item_column = column
                return item_column, item_row

    
def test():
    ''' '''
    files = ['cj_dlfx_1621401.xlsx', 'cj_dlfx_1621402.xlsx']
    item_name = "平时\n成绩"
    for file in files:
        item_column, item_row = get_item_row_column(file, item_name)
        print(item_column, item_row)


if __name__ == '__main__':
    ''' '''
    print('---Begin--------')
    test()
    print('---End--------')
