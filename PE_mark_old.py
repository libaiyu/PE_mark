#! python 3
# _*_ coding: utf_8  _*_

''' Modify  2017-6-24
         2017-7-11
'''


import os
import re
import openpyxl
import logging


MAXROW = 106
Reg = re.compile(r'标准表')

# TYPE_AGE = ['甲', '乙', '丙', '丁']
China_str = ['100米', '跳远', '铅球', '８００米', '男子成绩', '女子成绩']
PE_items = ['100米', '跳远', '铅球', '８００米']
TYPE = ['男子成绩', '女子成绩']

measure_information = []


def get_measure_information():
    '''先获取所需的信息：姓名，男女，项目，测量值，组别。'''
    wb = openpyxl.load_workbook('高考体育成绩测量表.xlsx')
    sheet = wb.get_active_sheet()
    col = 2
    for row in range(1, sheet.max_row+1):
        name = sheet.cell(row=row, column=col).value
        if name:
            type_age = sheet.cell(row=row, column=col-1).value
            gender = sheet.cell(row=row, column=col+1).value
            item_value = []
            for k in range(4):
                measure_value = sheet.cell(row=row, column=col+k+2).value
                item_value.append([PE_items[k], measure_value])
            measure_information.append((type_age, name, gender, item_value))
    for t in range(3):
        print(measure_information[t])
    print()


def information_to_mark(type_age, gender, item_name, measure_val):
    '''然后根据参数中提供的信息找到相应的成绩。'''
    if gender == '女':
        mark = mark_woman[(type_age, PE_items[item_name])][measure_val]
    elif gender == '男':
        mark = mark_man[(type_age, PE_items[item_name])][measure_val]
    return mark


# logging.basicConfig(level=logging.DEBUG,
# format='%(asctime)s-%(levelname)s-%(message)s')
logging.basicConfig(level=logging.ERROR,
                    format='%(asctime)s-%(levelname)s-%(message)s')
logging.critical('--------Start of program---------')

item_head = {}
mark_man = {}
mark_woman = {}
wb_val = {}

for file in os.listdir():
    if Reg.search(file):
        print(file)
        type_age = file[3]  # “标准表甲.xlsx”的第3个字符是“甲”。

        '''读成绩对照表，生成字典的字典。以供后续查询。 '''
        wb = openpyxl.load_workbook(file)
        sheet = wb.get_active_sheet()
        for col in range(1, sheet.max_column+1):
            key1 = col
            col_dict = {}
            for row in range(1, sheet.max_row+1):
                key2 = row
                val = sheet.cell(row=row, column=col).value   # 给定的列与行的值。
        # print(type(val))
        # input()
        # if val is not None:
        #     col_dict[key2] = val
                col_dict[key2] = val   # 字典－－遍历所有行。　｛行：值｝。
            wb_val[key1] = col_dict    # 字典－－遍历所有列。　{列：{行：值}}。
        print(key1-10, key2-100, wb_val[key1-10][key2-100])

        for col, row_val in wb_val.items():
            for row, val in row_val.items():
                if val in China_str:
                    if val in PE_items:   # 如果值是项目名。
                        item_head[val] = (col, row)   # 字典--项目名为键，所在的（列，行）为值。
        print(item_head)
        print()

        for val, (col, row) in item_head.items():
            key1 = (type_age, val)    # 组别，项目名
            item_man = {}
            item_woman = {}
            for r in range(row+2, MAXROW):
                key_m = wb_val[col][r]        # 男　的键　为　测量值
                val_m = wb_val[col+1][r]      # 男　的值　为　成绩
                item_man[key_m] = val_m       # 男　的字典　　测量值：成绩
                key_wo = wb_val[col+2][r]     # 女　的键　为　测量值
                val_wo = wb_val[col+3][r]     # 女　的值　为　成绩
                item_woman[key_wo] = val_wo   # 女　的字典
            mark_man[key1] = item_man         # 男　的字典的字典　项目名　为　键，（测量值：成绩）　为　值。
            mark_woman[key1] = item_woman     # 女　的字典的字典　项目名　为　键，（测量值：成绩）　为　值。

# print('男子成绩\n',mark_man,'\n')
# print('女子成绩\n',mark_woman,'\n')

'''读　‘高考体育成绩测量表.xlsx’，生成　字典的字典。'''
get_measure_information()

for each_information in measure_information:
    type_age = each_information[0]
#    print(each_information)
    gender = each_information[2]
    for k in range(4):
        item_name = k
        measure_val = each_information[3][k][1]
        mark = information_to_mark(type_age, gender, item_name, measure_val)
        each_information[3][k].append(mark)
    print(each_information)
#    input('for debug')
# for each_information in measure_information:
#    print(each_information)
''' write in a new file. '''
wb = openpyxl.Workbook()
sheet = wb.get_active_sheet()
row = 1
for each_information in measure_information:
    row += 1
    column = 2
    sheet.cell(row=row, column=column).value = each_information[0]
    sheet.cell(row=row, column=column+1).value = each_information[1]
    sheet.cell(row=row, column=column+2).value = each_information[2]
    sheet.cell(row=row, column=column+3).value = each_information[3][0][0]
    sheet.cell(row=row, column=column+4).value = each_information[3][0][1]
    sheet.cell(row=row, column=column+5).value = each_information[3][0][2]
    sheet.cell(row=row, column=column+6).value = each_information[3][1][0]
    sheet.cell(row=row, column=column+7).value = each_information[3][1][1]
    sheet.cell(row=row, column=column+8).value = each_information[3][1][2]
    sheet.cell(row=row, column=column+9).value = each_information[3][2][0]
    sheet.cell(row=row, column=column+10).value = each_information[3][2][1]
    sheet.cell(row=row, column=column+11).value = each_information[3][2][2]
    sheet.cell(row=row, column=column+12).value = each_information[3][3][0]
    sheet.cell(row=row, column=column+13).value = each_information[3][3][1]
    sheet.cell(row=row, column=column+14).value = each_information[3][3][2]
wb.save('体育成绩表3.xlsx')

logging.critical('-------End--------')
